"""
This program calculates the worked hours for a given week based on each day of the week.
The program must be able to receive hours as input and calculate the worked hours.
For example: Start Time --> 07:00 / Finish Time --> 17:00 / Total Worked Hours --> 10 Hours
The Program should be able to output floating point numbers, with up to 2 decimal places precision.

"""
from openpyxl import load_workbook


def calculate_worked_hours(start_time, finish_time):
	start_hour, start_minute = map(int, start_time.split(':'))
	finish_hour, finish_minute = map(int, finish_time.split(':'))

	total_minutes = (finish_hour - start_hour) * 60 + (finish_minute - start_minute)
	total_hours = total_minutes / 60

	return round(total_hours, 2)


class Calculator:
	def __init__(self):
		self.week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
		self.worked_hours_dict = {}
		self.current_row = 2
		self.workbook = load_workbook(filename='calculator.xlsx')
		self.sheet = self.workbook.active
		self.current_month = None
		self.starting_day = None

	def send_to_sheet(self):
		for day in self.week_days:
			print(f"--> {day.upper()} <-- Press Enter to Skip")
			start_hours = input(f'In: ')
			finish_hours = input(f'Out: ')

			# Assign the values to the specific cells
			# Assuming Day is Column 1, Weekday is Column 2 and Hours is Column 3
			self.sheet.cell(row=self.current_row, column=1, value=f'{self.starting_day} {self.current_month}')
			self.sheet.cell(row=self.current_row, column=2, value=day)

			if start_hours and finish_hours:
				self.worked_hours_dict[day] = calculate_worked_hours(start_hours, finish_hours)
				self.sheet.cell(row=self.current_row, column=3, value=self.worked_hours_dict[day])
			else:
				self.sheet.cell(row=self.current_row, column=3, value='')

			# Move to next row for next iteration
			self.starting_day += 1
			self.current_row += 1

		# Save the workbook after all updates

		self.workbook.save(filename='calculator.xlsx')

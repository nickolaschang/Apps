import tkinter as tk
from openpyxl import load_workbook


class Calculator:
	def __init__(self):
		self.week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
		self.worked_hours_dict = {}
		self.current_row = 2
		self.workbook = load_workbook(filename='calculator.xlsx')
		self.sheet = self.workbook.active
		self.current_month = None
		self.starting_day = None

	@staticmethod
	def calculate_worked_hours(start_time, finish_time):
		start_hour, start_minute = map(int, start_time.split(':'))
		finish_hour, finish_minute = map(int, finish_time.split(':'))
		total_minutes = (finish_hour - start_hour) * 60 + (finish_minute - start_minute)
		total_hours = total_minutes / 60
		return round(total_hours, 2)

	def send_to_sheet(self):
		for day in self.week_days:
			self.sheet.cell(row=self.current_row, column=1, value=f'{self.starting_day} {self.current_month}')
			self.sheet.cell(row=self.current_row, column=2, value=day)
			if day in self.worked_hours_dict:
				self.sheet.cell(row=self.current_row, column=3, value=self.worked_hours_dict[day])
			else:
				self.sheet.cell(row=self.current_row, column=3, value='')
			self.starting_day += 1
			self.current_row += 1
		self.workbook.save(filename='calculator.xlsx')


class App:
	def __init__(self, window, calculator):
		self.window = window
		self.calculator = calculator
		self.entries_in = {}
		self.entries_out = {}
		self.entry_month = tk.Entry(window)
		self.entry_start_day = tk.Entry(window)
		self.entry_month.grid(row=0, column=4)
		self.entry_start_day.grid(row=1, column=4)
		self.window.title('Hours Calculator')

		for i in range(5):  # for 5 columns only, increase or adjust the range for better modularization?
			self.window.columnconfigure(i, weight=1)

		self.window.geometry('800x250')

		tk.Label(window, text='Day').grid(column=0, row=0)
		tk.Label(window, text='In').grid(column=1, row=0)
		tk.Label(window, text='Out').grid(column=2, row=0)
		tk.Label(window, text="Month:").grid(column=3, row=0)
		tk.Label(window, text="Start day:").grid(column=3, row=1)

		for i, day in enumerate(self.calculator.week_days, start=1):
			tk.Label(window, text=day).grid(column=0, row=i)
			self.entries_in[day] = tk.Entry(window)
			self.entries_out[day] = tk.Entry(window)
			self.entries_in[day].grid(column=1, row=i)
			self.entries_out[day].grid(column=2, row=i)

		self.message_log = tk.Listbox(window)
		self.message_log.grid(column=4, row=2, rowspan=len(self.calculator.week_days), columnspan=3, sticky='ew')

		submit = tk.Button(window, text="Submit", command=self.calculate_hours)
		submit.grid(column=0, row=i + 1, columnspan=3)

	def calculate_hours(self):
		try:
			self.calculator.current_month = self.entry_month.get()
			self.calculator.starting_day = int(self.entry_start_day.get())
			for day in self.calculator.week_days:
				start = self.entries_in[day].get()
				end = self.entries_out[day].get()
				if start and end:  # if both fields are not empty
					self.calculator.worked_hours_dict[day] = \
						self.calculator.calculate_worked_hours(start, end)
			self.calculator.send_to_sheet()
			self.message_log.insert(tk.END, 'File Saved!')
		except ValueError:
			self.message_log.insert(tk.END, "Start Day must be a number!!")


if __name__ == '__main__':
	window = tk.Tk()
	calc = Calculator()
	app = App(window, calc)
	window.mainloop()
